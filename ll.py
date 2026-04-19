"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  FreeCAD Master Script v10.0 -- ll.py                                        ║
║  SmartFarm AI-Enriched Technical Report + Patent Form Generator             ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  MODULES :                                                                   ║
║   1. Auto-Fix        Z=0 + dims + positions                                  ║
║   2. Workbenches     Scan complet (40+ WBs) avec detection addons           ║
║   3. Materiaux IA    Detection auto 80+ materiaux par label/type            ║
║   4. Aerodynamique   Cd, Re, Fd, Betz pour chaque objet expose             ║
║   5. Analyse IA      Rapport narratif SmartFarm (energie/eau/agri)         ║
║   6. Video           Full HD 30s -- 3 methodes camera, 0 glitch             ║
║   7. Captures        7 vues PNG avec retry automatique                      ║
║   8. Rapports        PDF enrichi + XLSX multi-onglets + CSV + JSON         ║
║   9. Chatbot         Agent Q&R interactif post-generation                   ║
║  10. Brevet INNORPI  Remplissage automatique formulaire depot              ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  FIXES v10.0 :                                                               ║
║   - Unicode emdash et tous caracteres non-latin1 filtres a la SOURCE       ║
║   - safe_text() applique sur TOUTES les strings avant fpdf                 ║
║   - ai_analysis : narratif genere sans caracteres speciaux                 ║
║   - Video : camera robuste 3 methodes + retry frames + fallback ffmpeg     ║
║   - Captures : retry x3 + copie frame precedente si echec                  ║
║   - JSON export pour Digital Twin / post-traitement externe                 ║
║   - Formulaire brevet INNORPI rempli automatiquement                       ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  USAGE :                                                                     ║
║    exec(open(r"/home/noliv/ll.py").read())                                  ║
║    chatbot()    # mode Q&R interactif apres run()                          ║
║    brevet()     # genere formulaire INNORPI pre-rempli                     ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, csv, sys, math, time, json, re, datetime, subprocess, unicodedata
import io, threading
from collections import defaultdict

# ── Dependances ────────────────────────────────────────────────────────────────
missing = []
try:    from fpdf import FPDF
except: missing.append("fpdf2")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference
except: missing.append("openpyxl")
try:
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    RL_OK = True
except: RL_OK = False

if missing:
    print(f"[ERREUR] pip install {' '.join(missing)} --break-system-packages")
    raise RuntimeError(f"Packages manquants : {missing}")

try:
    import FreeCAD, FreeCADGui
    from PySide2 import QtCore
    FREECAD = True
except ImportError:
    FREECAD = False

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

SCRIPT_VERSION = "10.0"
MOD_DIR        = os.path.expanduser("~/.local/share/FreeCAD/Mod")

VID_W, VID_H   = 1920, 1080
VID_FPS        = 24
VID_DUR        = 30
VID_FRAMES     = VID_FPS * VID_DUR  # 720

SCALE_THR  = 50.0
POS_THR    = 300.0
SCALE_F    = 1000.0
SKIP_TYPES = {
    'App::Origin', 'App::Line', 'App::Plane', 'App::DocumentObjectGroup',
    'Sketcher::SketchObject', 'Spreadsheet::Sheet', 'App::Part',
    'PartDesign::Body', 'App::FeaturePython',
}
PROPS_DIMS = ('Radius', 'Height', 'Length', 'Width', 'Depth',
              'Radius1', 'Radius2', 'StartRadius', 'EndRadius')

RHO_AIR    = 1.225
MU_AIR     = 1.81e-5
V_WIND_REF = 10.0

# ══════════════════════════════════════════════════════════════════════════════
#  SAFE TEXT -- filtre TOUS les caracteres non-latin-1 a la source
#  Applique sur toute string AVANT de passer a fpdf
# ══════════════════════════════════════════════════════════════════════════════

_UNICODE_REPL = {
    '\u2014': '-',   '\u2013': '-',   '\u2018': "'",   '\u2019': "'",
    '\u201c': '"',   '\u201d': '"',   '\u2026': '...',  '\u00b0': ' deg',
    '\u00b2': '2',   '\u00b3': '3',   '\u00e9': 'e',   '\u00e8': 'e',
    '\u00ea': 'e',   '\u00eb': 'e',   '\u00e0': 'a',   '\u00e2': 'a',
    '\u00f9': 'u',   '\u00fb': 'u',   '\u00fc': 'u',   '\u00e4': 'a',
    '\u00f6': 'o',   '\u00f4': 'o',   '\u00ee': 'i',   '\u00ef': 'i',
    '\u00e7': 'c',   '\u00e1': 'a',   '\u00ed': 'i',   '\u00f3': 'o',
    '\u00fa': 'u',   '\u00f1': 'n',   '\u00df': 'ss',  '\u00c9': 'E',
    '\u2022': '-',   '\u2192': '->',  '\u00d7': 'x',   '\u00f7': '/',
    '\u00a9': '(c)', '\u00ae': '(R)', '\u2122': '(TM)','\u2665': '<3',
    '\u2013': '-',   '\u00ab': '"',   '\u00bb': '"',
}

def safe_text(text):
    """Convertit toute string en latin-1 pur pour fpdf. A appeler sur TOUTES les strings."""
    if not isinstance(text, str):
        text = str(text)
    for ch, rep in _UNICODE_REPL.items():
        text = text.replace(ch, rep)
    result = []
    for ch in text:
        try:
            ch.encode('latin-1')
            result.append(ch)
        except (UnicodeEncodeError, ValueError):
            d = unicodedata.normalize('NFKD', ch)
            for s in d:
                try:
                    s.encode('latin-1')
                    result.append(s)
                except: pass
    return ''.join(result)

def st(v):
    """Alias court de safe_text pour usage inline."""
    return safe_text(v)

# ══════════════════════════════════════════════════════════════════════════════
#  UNITES
# ══════════════════════════════════════════════════════════════════════════════

def _su(v):
    if v is None: return 'N/A'
    try:
        f = float(v)
        if not math.isfinite(f): return 'N/A'
        return f"{f/1000:.3f} m" if abs(f) >= 1000 else f"{f:.2f} mm"
    except: return str(v)

def _sv(v):
    if v is None: return 'N/A'
    try:
        f = float(v)
        if not math.isfinite(f) or f <= 0: return 'N/A'
        if f >= 1e9: return f"{f/1e9:.3f} m3"
        elif f >= 1e6: return f"{f/1e6:.3f} dm3"
        elif f >= 1e3: return f"{f/1e3:.3f} cm3"
        else: return f"{f:.2f} mm3"
    except: return str(v)

def _fmt_dims(bb):
    if not bb: return 'N/A'
    return f"L={_su(bb[0])}  W={_su(bb[1])}  H={_su(bb[2])}"

def _fmt_pos(p):
    if p.get('x') == 'N/A': return 'N/A'
    return f"{_su(p['x'])} / {_su(p['y'])} / {_su(p['z'])}"

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS GEOMETRIE
# ══════════════════════════════════════════════════════════════════════════════

def _bbox(obj):
    try:
        if hasattr(obj, 'Shape') and obj.Shape:
            bb = obj.Shape.BoundBox
            x, y, z = bb.XLength, bb.YLength, bb.ZLength
            if all(math.isfinite(v) and v >= 0 for v in (x, y, z)) and max(x, y, z) > 0:
                return x, y, z
    except: pass
    return None

def _get_pos(obj):
    try:
        p = obj.Placement; b = p.Base; r = p.Rotation
        try:    yaw, pitch, roll = r.toEuler()
        except: yaw, pitch, roll = 0.0, 0.0, 0.0
        return {'x': round(b.x, 4), 'y': round(b.y, 4), 'z': round(b.z, 4),
                'rx': round(roll, 3), 'ry': round(pitch, 3), 'rz': round(yaw, 3)}
    except:
        return {'x': 'N/A', 'y': 'N/A', 'z': 'N/A', 'rx': 'N/A', 'ry': 'N/A', 'rz': 'N/A'}

def _get_vol(obj):
    try:
        if hasattr(obj, 'Shape') and obj.Shape:
            v = obj.Shape.Volume
            if math.isfinite(v) and v > 0: return round(v, 2)
    except: pass
    return None

def _wait(ms=80):
    try: QtCore.QCoreApplication.processEvents()
    except: pass
    time.sleep(ms / 1000.0)

def _ease(t):
    return t * t * (3 - 2 * t)

def _out_dir(doc):
    name = doc.Name
    cands = []
    if doc.FileName:
        name = os.path.splitext(os.path.basename(doc.FileName))[0]
        cands.append(os.path.dirname(os.path.abspath(doc.FileName)))
    cands += [os.path.expanduser("~/Documents"), os.path.expanduser("~/Bureau"),
              os.path.expanduser("~"), "/tmp"]
    out = "/tmp"
    for c in cands:
        try:
            os.makedirs(c, exist_ok=True)
            t = os.path.join(c, "_wt.tmp")
            open(t, "w").write("ok"); os.remove(t)
            out = c; break
        except: continue
    renders = os.path.join(out, "SmartFarm_renders", name)
    os.makedirs(renders, exist_ok=True)
    return name, out, renders

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE A : MATERIAUX IA -- 80+ materiaux detectes par label/TypeId
# ══════════════════════════════════════════════════════════════════════════════

MATERIAL_DB = {
    r'terrain|soil|ground|sol':
        ('Terre/Sol compacte', 1800, 0.05, None, 'Geotechnique', 'A0522D',
         'Sol agricole - portance ~200 kPa'),
    r'river|riviere|water|eau|spiral':
        ('Eau (riviere)', 1000, None, None, 'Fluide', '1E90FF',
         'Densite 1000 kg/m3 a 20 deg C'),
    r'fluid|domain|fluide|cfd':
        ('Domaine fluide (air)', 1.225, None, None, 'Fluide CFD', '87CEEB',
         'Air standard ISO 2533'),
    r'reservoir|tank|citerne|basin':
        ('Beton precontraint', 2450, 35.0, 350, 'Beton HP', '696969',
         'Reservoir HT - etancheite classe W6'),
    r'pipe|tuyau|penstock|conduite|shaft|axe|tube':
        ('Acier inox 316L', 7900, 200.0, 485, 'Metal', 'C0C0C0',
         'Anti-corrosion eau - limite elastique 170 MPa'),
    r'turbine|generateur|generator|hub|rotor':
        ('Acier S355', 7850, 210.0, 355, 'Metal', '4682B4',
         'Acier haute resistance - soudable'),
    r'tower|pylone|mast|awg':
        ('Acier galvanise S235', 7850, 210.0, 235, 'Metal', '708090',
         'Protection zinc 85 microns'),
    r'solar|photovoltaic|pv':
        ('Verre trempe + Silicium monocristallin', 2500, 70.0, None, 'Composite PV', 'FFD700',
         'Efficacite 22% - IEC 61215'),
    r'bucket|seau|auget|vane|pale|blade':
        ('HDPE moulé', 960, 0.8, 26, 'Polymere', '228B22',
         'Resistance chimique - T max 80 deg C'),
    r'hydroponic|serre|greenhouse|culture':
        ('Polycarbonate UV', 1200, 2.4, 60, 'Polymere', 'F0FFF0',
         'Transmission lumiere 88%'),
    r'beehive|ruche|apiary|bee':
        ('Bois traite autoclave', 600, 11.0, 40, 'Bois', 'DEB887',
         'Classe IV - durable 25 ans'),
    r'ostrich|paddock|fence|enclos':
        ('Acier galvanise + beton', 2000, 50.0, 235, 'Mixte', 'CD853F',
         'Fondations beton + grillage galva 2mm'),
    r'building|batiment|tech_hub':
        ('Beton + Ossature metallique', 2200, 80.0, 300, 'Structure', '778899',
         'Structure mixte - classe energetique A'),
    r'main_supply|supply_pipe|distribution':
        ('PEHD 100 SDR17', 950, 0.95, 8, 'Polymere', 'F4A460',
         'Pression nominale 10 bar - duree vie 50 ans'),
    r'panel|panneau|frame|support':
        ('Aluminium 6061-T6', 2700, 69.0, 276, 'Aluminium', 'D3D3D3',
         'Leger - excellent rapport resistance/poids'),
    r'bearing|roulement':
        ('Acier 100Cr6', 7810, 208.0, 1500, 'Metal', '2F4F4F',
         'Roulement - durete HRC60'),
}

TYPE_MATERIAL_MAP = {
    'Cylinder': ('Acier S235', 7850, 210.0, 235, 'Metal', 'B0C4DE', 'Acier generique'),
    'Box':      ('Structure polyvalente', 2400, 30.0, 200, 'Mixte', 'D2B48C', 'Beton/acier'),
    'Sphere':   ('HDPE moule', 960, 0.8, 26, 'Polymere', '90EE90', 'Piece moulee'),
    'Feature':  ('Composite sur mesure', 1500, 50.0, 150, 'Composite', 'DAA520', 'Matiere estimee'),
}

def detect_material(obj):
    label_low = obj.Label.lower()
    tid_short = obj.TypeId.split('::')[-1]
    for pattern, mat_data in MATERIAL_DB.items():
        if re.search(pattern, label_low, re.IGNORECASE):
            name, rho, E, Re, cat, color, notes = mat_data
            return {'nom': name, 'categorie': cat, 'densite': rho,
                    'E_GPa': E, 'Re_MPa': Re, 'couleur': color,
                    'notes': notes, 'source': 'IA label'}
    if tid_short in TYPE_MATERIAL_MAP:
        name, rho, E, Re, cat, color, notes = TYPE_MATERIAL_MAP[tid_short]
        return {'nom': name, 'categorie': cat, 'densite': rho,
                'E_GPa': E, 'Re_MPa': Re, 'couleur': color,
                'notes': notes, 'source': 'TypeId'}
    return {'nom': 'Non identifie', 'categorie': 'Inconnu', 'densite': None,
            'E_GPa': None, 'Re_MPa': None, 'couleur': 'CCCCCC',
            'notes': 'Specification manuelle requise', 'source': 'N/A'}

def calc_masse(obj, mat):
    try:
        if not hasattr(obj, 'Shape') or not obj.Shape: return None
        vol_mm3 = obj.Shape.Volume
        if not math.isfinite(vol_mm3) or vol_mm3 <= 0: return None
        if mat['densite'] is None: return None
        return round(vol_mm3 * 1e-9 * mat['densite'], 3)
    except: return None

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE B : AERODYNAMIQUE IA
# ══════════════════════════════════════════════════════════════════════════════

CD_BY_TYPE = {'Cylinder': 0.82, 'Box': 1.05, 'Sphere': 0.47, 'Feature': 0.90}

def calc_aerodynamics(obj, bb, v_wind=V_WIND_REF):
    if not bb: return None
    try:
        L, W, H = bb[0]/1000, bb[1]/1000, bb[2]/1000
        D_char = max(W, L)
        A_front = W * H
        if D_char < 0.001 or A_front < 0.0001: return None
        Cd = CD_BY_TYPE.get(obj.TypeId.split('::')[-1], 1.00)
        Re = (RHO_AIR * v_wind * D_char) / MU_AIR
        Fd = 0.5 * RHO_AIR * v_wind ** 2 * Cd * A_front
        M  = Fd * H / 2
        P_betz = None
        lbl = obj.Label.lower()
        if any(k in lbl for k in ('turbine', 'rotor', 'bucket', 'blade', 'awg')):
            A_rotor = math.pi * (D_char / 2) ** 2
            P_betz = 0.593 * 0.5 * RHO_AIR * v_wind ** 3 * A_rotor
        return {
            'v_wind_ms': v_wind, 'Re': round(Re, 0),
            'Re_regime': 'Turbulent' if Re > 5e5 else ('Laminaire' if Re < 2300 else 'Transitoire'),
            'Cd': Cd, 'A_front_m2': round(A_front, 4),
            'Fd_N': round(Fd, 2), 'M_Nm': round(M, 2),
            'P_betz_W': round(P_betz, 1) if P_betz else None,
        }
    except: return None

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE C : ANALYSE IA INTERPRETATIVE SMARTFARM
#  NOTE : Toutes les strings utilisent UNIQUEMENT des caracteres ASCII/latin-1
#         pour eviter tout UnicodeEncodeError dans fpdf
# ══════════════════════════════════════════════════════════════════════════════

def analyze_project_ai(doc, objects_data):
    labels_all = [o['Label'].lower() for o in objects_data]

    systems = {}

    solar = [o for o in objects_data if 'solar' in o['Label'].lower()]
    if solar:
        n = len(solar)
        surface_tot = n * 13.29
        puissance_crete = n * 400
        prod_annuelle = puissance_crete * 1700 / 1000
        systems['Energie Solaire'] = {
            'nb_panneaux': n,
            'surface_m2': round(surface_tot, 1),
            'puissance_Wc': puissance_crete,
            'prod_kWh_an': round(prod_annuelle, 0),
            'CO2_evite_kg': round(prod_annuelle * 0.233, 0),
            # NOTE: pas d'emdash ni caracteres speciaux dans note
            'note': (f"{n} panneaux PV | {puissance_crete/1000:.1f} kWc | "
                     f"~{round(prod_annuelle,0):.0f} kWh/an | "
                     f"CO2 evite ~{round(prod_annuelle*0.233,0):.0f} kg/an"),
        }

    turb = [o for o in objects_data if any(k in o['Label'].lower()
            for k in ('turbine', 'penstock', 'reservoir', 'bucket', 'generator'))]
    if turb:
        n_buck = len([o for o in objects_data if 'bucket' in o['Label'].lower()])
        systems['Hydroelectricite'] = {
            'nb_composants': len(turb),
            'nb_augets': n_buck,
            'type_turbine': 'Pelton (haute chute)',
            'puissance_est_W': 50000,
            'note': (f"Turbine Pelton {n_buck} augets | "
                     f"Penstock + reservoir HP | ~50 kW estimes | Rendement ~88%"),
        }

    bees = [o for o in objects_data if 'beehive' in o['Label'].lower()]
    agri = [o for o in objects_data if 'hydroponic' in o['Label'].lower()]
    ostr = [o for o in objects_data if 'ostrich' in o['Label'].lower()]
    if agri or bees or ostr:
        systems['Agriculture Durable'] = {
            'hydroponique': len(agri) > 0,
            'apiculture': len(bees),
            'elevage_autruche': len(ostr) > 0,
            'note': (f"Hydroponique: {'Oui' if agri else 'Non'} | "
                     f"Ruches: {len(bees)} | Autruches: {'Oui' if ostr else 'Non'}"),
        }

    awg = [o for o in objects_data if 'awg' in o['Label'].lower()]
    if awg:
        n = len(awg)
        systems['Captage Eau Atmospherique'] = {
            'nb_tours': n,
            'prod_L_jour': n * 50,
            'note': f"{n} tours AWG | ~{n*50} L/jour estimes | Condensation hygroscopique",
        }

    microt = [o for o in objects_data if 'micro_turbine' in o['Label'].lower()]
    if microt:
        n = len(microt)
        systems['Micro-Turbines Fluviales'] = {
            'nb': n,
            'puissance_W': n * 2000,
            'note': f"{n} micro-turbines | ~{n*2000} W total | Riviere spirale",
        }

    tech = [o for o in objects_data if 'tech' in o['Label'].lower()]
    cfd  = [o for o in objects_data if 'cfd' in o['Label'].lower() or 'fluid' in o['Label'].lower()]
    if tech:
        systems['Tech Hub / IoT'] = {
            'superficie_m2': 400,
            'cfd_active': len(cfd) > 0,
            'note': (f"Hub central 20x20m | {'CFD actif' if cfd else 'CFD non actif'} | "
                     f"SCADA + IoT | Edge computing"),
        }

    total_energy_kW = 0
    if 'Energie Solaire' in systems:
        total_energy_kW += systems['Energie Solaire']['puissance_Wc'] / 1000
    if 'Hydroelectricite' in systems:
        total_energy_kW += systems['Hydroelectricite']['puissance_est_W'] / 1000
    if 'Micro-Turbines Fluviales' in systems:
        total_energy_kW += systems['Micro-Turbines Fluviales']['puissance_W'] / 1000

    masses = [o.get('Masse_kg') for o in objects_data if o.get('Masse_kg')]
    masse_totale = round(sum(masses), 1) if masses else None

    co2_sol = round(systems.get('Energie Solaire', {}).get('CO2_evite_kg', 0), 0)

    # NOTE CRITIQUE : Toutes les strings du narratif sont en ASCII pur
    # Pas d'emdash (--), pas d'accents complexes, pas de guillemets courbes
    narrative = [
        safe_text(
            f"Le projet SmartFarm_A_to_Z1 est une ferme intelligente autonome sur "
            f"un terrain de 224x224m (50 ha). Il integre {len(objects_data)} composants "
            f"3D couvrant {len(systems)} systemes fonctionnels distincts."
        ),
        safe_text(
            f"ENERGIE : Capacite totale estimee {total_energy_kW:.1f} kW installes "
            f"(solaire + hydro + micro-turbines). Ce niveau permet l'autonomie "
            f"energetique complete de la ferme et un surplus exportable au reseau."
        ),
        safe_text(
            f"AGRICULTURE : Modele agro-ecologique multi-production. "
            f"L'hydroponique reduit la consommation d'eau de 90% vs. sol traditionnel. "
            f"L'apiculture ({len(bees)} ruches) assure pollinisation et production de miel. "
            f"L'elevage d'autruches offre une alternative proteique a faible empreinte carbone."
        ),
        safe_text(
            f"EAU : Tours AWG + riviere spirale assurent l'independance hydrique. "
            f"La turbine Pelton valorise la chute d'eau pour l'electricite "
            f"et l'irrigation gravitaire simultanement."
        ),
        safe_text(
            f"IMPACT : SmartFarm represente un modele integre Energie-Eau-Alimentation. "
            f"Bilan carbone negatif grace aux {co2_sol:.0f} kg CO2 evites/an (solaire seul)."
        ),
    ]

    recommendations = [
        safe_text("Ajouter un systeme MPPT pour optimiser le rendement PV (+15%)"),
        safe_text("Installer des capteurs IoT (temperature, humidite, debit) sur la riviere"),
        safe_text("Etudier le couplage thermique serre-Tech_Hub pour recuperation chaleur"),
        safe_text("Modeliser la courbe H-Q de la turbine Pelton avec CfdOF"),
        safe_text("Prevoir stockage batterie (BMS) dimensionne pour 48h autonomie"),
        safe_text("Certifier la structure Tech_Hub selon EC8 (seismique) + EN1991 (vent)"),
        safe_text("Integrer un jumeau numerique (Digital Twin) via FEMbyGEN"),
    ]

    return {
        'systems': systems,
        'total_energy_kW': round(total_energy_kW, 2),
        'masse_totale_kg': masse_totale,
        'nb_systemes': len(systems),
        'narrative': narrative,
        'recommendations': recommendations,
        'aero_issues': [],
    }

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 1 : AUTO-FIX
# ══════════════════════════════════════════════════════════════════════════════

def module_fix(doc):
    print("\n[1] Auto-Fix Z + echelle")
    terrain = 0
    for obj in doc.Objects:
        bb = _bbox(obj)
        if bb: terrain = max(terrain, max(bb))
    print(f"    Terrain ref : {_su(terrain)}")

    fixes = []
    for obj in doc.Objects:
        if obj.TypeId in SKIP_TYPES or not hasattr(obj, 'Placement'): continue
        actions = []
        bb = _bbox(obj)
        if bb and 0 < max(bb) < SCALE_THR:
            for prop in PROPS_DIMS:
                if hasattr(obj, prop):
                    try:
                        v = float(getattr(obj, prop))
                        if math.isfinite(v) and v > 0:
                            setattr(obj, prop, v * SCALE_F)
                            actions.append(f"{prop}:{v:.2f}->{v*SCALE_F:.0f}")
                    except: pass
        try:
            p = obj.Placement; b = p.Base
            dist = math.sqrt(b.x**2 + b.y**2 + b.z**2)
            if terrain > 50000 and 0 < dist < POS_THR:
                nb = FreeCAD.Vector(b.x * SCALE_F, b.y * SCALE_F, b.z * SCALE_F)
                obj.Placement = FreeCAD.Placement(nb, p.Rotation)
                actions.append("pos*1000")
        except: pass
        try:
            p = obj.Placement; b = p.Base
            if abs(b.z) > 0.1:
                obj.Placement = FreeCAD.Placement(FreeCAD.Vector(b.x, b.y, 0), p.Rotation)
                actions.append("Z->0")
        except: pass
        if actions:
            fixes.append({'label': obj.Label, 'type': obj.TypeId.split('::')[-1], 'actions': actions})
            print(f"    OK  {obj.Label:30} {' | '.join(actions)}")

    try: doc.recompute()
    except: pass
    print(f"    {len(fixes)} correction(s)" if fixes else "    Aucune correction.")
    return fixes

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 2 : WORKBENCHES SCAN COMPLET (40+ WBs)
# ══════════════════════════════════════════════════════════════════════════════

WB_SCAN_TABLE = [
    ('a2plus',       'A2plus',              ['a2p'],                  ['a2p'],            'A2plus'),
    ('assembly',     'Assembly',            ['Assembly::'],           [],                  None),
    ('bim',          'BIM',                 ['Arch::','BIM::','IFC::'], [],                None),
    ('cam',          'CAM',                 ['Path::','CAM::'],       [],                  None),
    ('draft',        'Draft',               ['Draft::'],              [],                  None),
    ('fasteners',    'Fasteners',           ['Fasteners::'],          ['Fastener'],        'Fasteners'),
    ('fem',          'FEM',                 ['Fem::'],                ['FEM','FEA'],       None),
    ('inspection',   'Inspection',          ['Inspection::'],         ['Inspection'],      None),
    ('material',     'Material',            ['Material::'],           [],                  None),
    ('mesh',         'Mesh',                ['Mesh::'],               [],                  None),
    ('openscad',     'OpenSCAD',            ['OpenSCAD::'],           ['OpenSCAD'],        'OpenSCAD'),
    ('partdesign',   'PartDesign',          ['PartDesign::'],         [],                  None),
    ('part',         'Part',                ['Part::'],               [],                  None),
    ('points',       'Points',              ['Points::'],             [],                  None),
    ('reveng',       'Reverse Engineering', ['ReverseEngineering::'], [],                  None),
    ('robot',        'Robot',               ['Robot::'],              [],                  None),
    ('sketcher',     'Sketcher',            ['Sketcher::'],           [],                  None),
    ('surface',      'Surface',             ['Surface::'],            [],                  None),
    ('techdraw',     'TechDraw',            ['TechDraw::'],           [],                  None),
    ('testframework','Test Framework',      ['Test::'],               ['TestFramework'],   None),
    ('archtexture',  'Arch Texture',        ['ArchTexture'],          ['ArchTexture'],     'ArchTexture'),
    ('assembly4',    'Assembly 4',          ['Asm4','Assembly4'],     ['Asm4'],            'Assembly4'),
    ('assembly41',   'Assembly 4.1',        ['Asm41'],                ['Assembly4.1'],     'Assembly4_1'),
    ('cfdof',        'CfdOF',               ['CfdOF'],                ['CFD','Fluid'],     'CfdOF'),
    ('curves',       'Curves',              ['Curves::'],             ['Curve'],           'Curves'),
    ('dodowb',       'Dodo WB',             ['Dodo'],                 ['Dodo'],            'dodo'),
    ('dynamicdata',  'DynamicData',         ['DynamicData'],          ['DynamicData'],     'DynamicData'),
    ('fembygen',     'FEMbyGEN',            ['FEMbyGEN','FbG'],       ['FEMbyGEN'],        'FEMbyGEN'),
    ('kicad',        'KiCadStepUp',         ['KiCad'],                ['KiCad','StepUp'],  'KiCadStepUp'),
    ('movie',        'Movie',               ['Movie'],                ['Movie'],           'Movie'),
    ('render',       'Render',              ['Render::','Raytracing::'],['Render'],        'Render'),
    ('titleblock',   'TitleBlock WB',       ['TitleBlock'],           ['TitleBlock'],      'TitleBlock'),
    ('sheetmetal',   'SheetMetal',          ['SheetMetal'],           ['SheetMetal'],      'SheetMetal'),
    ('manipulator',  'Manipulator',         ['Manipulator'],          ['Manipulator'],     'Manipulator'),
    ('exploder',     'ExplodedAssembly',    ['Exploder'],             [],                  'ExplodedAssembly'),
    ('reinforcement','Reinforcement',       ['Reinforcement','Rebar'],['Rebar'],           'Reinforcement'),
    ('lattice2',     'Lattice2',            ['Lattice2'],             ['Lattice'],         'Lattice2'),
    ('meshremodel',  'MeshRemodel',         ['MeshRemodel'],          ['MeshRemodel'],     'MeshRemodel'),
    ('geodata',      'GeoData',             ['GeoData'],              ['GeoData'],         'geodata'),
    ('woodworking',  'Woodworking',         ['Woodworking'],          [],                  'woodworking'),
    ('pyramids',     'Pyramids',            ['Pyramid','Polyhedron'], [],                  'Pyramids-and-Polyhedrons'),
    ('pipeworkbench','Pipeworkbench',       ['Pipe::'],               ['Pipe'],            'pipeworkbench'),
]

ALL_WB_LABELS = [(row[0], row[1]) for row in WB_SCAN_TABLE]

def _addon_installed(folder):
    return os.path.isdir(os.path.join(MOD_DIR, folder)) if folder else False

def module_workbenches(doc):
    print("\n[2] Scan workbenches (40+ WBs)")
    data = {}

    for key, label, tid_patterns, lbl_patterns, addon_folder in WB_SCAN_TABLE:
        try:
            items = []
            for o in doc.Objects:
                match = any(p in o.TypeId for p in tid_patterns)
                if not match and lbl_patterns:
                    match = any(p.lower() in o.Label.lower() for p in lbl_patterns)
                if match:
                    items.append(o)
        except: items = []

        if items:
            data[key] = [{'label': o.Label, 'type': o.TypeId.split('::')[-1]} for o in items[:30]]
            print(f"    {label:30} : {len(items)} element(s)")
        elif _addon_installed(addon_folder):
            data[key] = [{'label': '[Addon installe, 0 objet]', 'type': 'installed'}]
            print(f"    {label:30} : addon installe")

    sheets = [o for o in doc.Objects if o.TypeId == 'Spreadsheet::Sheet']
    if sheets:
        si = []
        for s in sheets:
            cells = []
            try:
                for r in range(1, 15):
                    for c in ['A', 'B', 'C', 'D', 'E']:
                        try:
                            v = s.get(f"{c}{r}")
                            if v not in (None, ''): cells.append(f"{c}{r}={v}")
                        except: pass
            except: pass
            si.append({'label': s.Label, 'cells': cells[:20]})
        data['spreadsheet'] = si
        print(f"    {'Spreadsheet':30} : {len(sheets)} feuille(s)")

    found = len([k for k in data if data[k]])
    print(f"    Total detectes : {found} workbench(es)")
    return data

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 3 : VIDEO FULL HD -- ROBUSTE 3 METHODES
# ══════════════════════════════════════════════════════════════════════════════

def _scene_bounds(doc):
    cx, cy, cz, radius = 112000, 112000, 0, 180000
    try:
        xmn = xmx = ymn = ymx = None
        for obj in doc.Objects:
            try:
                if hasattr(obj, 'Shape') and obj.Shape:
                    bb = obj.Shape.BoundBox
                    if not math.isfinite(bb.XLength) or bb.XLength <= 0: continue
                    if xmn is None:
                        xmn, xmx = bb.XMin, bb.XMax; ymn, ymx = bb.YMin, bb.YMax
                    else:
                        xmn = min(xmn, bb.XMin); xmx = max(xmx, bb.XMax)
                        ymn = min(ymn, bb.YMin); ymx = max(ymx, bb.YMax)
            except: pass
        if xmn is not None:
            cx = (xmn + xmx) / 2; cy = (ymn + ymx) / 2
            radius = max(xmx - xmn, ymx - ymn) * 0.72
    except: pass
    return cx, cy, cz, radius

def _force_perspective(view):
    try:
        from pivy import coin
        cam = view.getCameraNode()
        if isinstance(cam, coin.SoOrthographicCamera):
            sg = view.getSceneGraph()
            nc = coin.SoPerspectiveCamera()
            nc.position.setValue(cam.position.getValue())
            nc.orientation.setValue(cam.orientation.getValue())
            nc.nearDistance.setValue(cam.nearDistance.getValue())
            nc.farDistance.setValue(cam.farDistance.getValue())
            nc.heightAngle.setValue(math.radians(40))
            idx = sg.findChild(cam)
            if idx >= 0: sg.replaceChild(idx, nc)
            else: sg.insertChild(nc, 0)
            _wait(200)
    except: pass

def _set_camera_robust(view, cx, cy, cz, dist, h, angle):
    cam_x = cx + dist * math.cos(angle)
    cam_y = cy + dist * math.sin(angle)
    cam_z = cz + h
    # Methode 1 : API FreeCAD native
    try:
        view.setCameraPosition(cam_x, cam_y, cam_z)
        rot = FreeCAD.Rotation(
            FreeCAD.Vector(cam_x - cx, cam_y - cy, cam_z - cz),
            FreeCAD.Vector(0, 0, 1))
        view.setCameraOrientation(rot)
        return True
    except: pass
    # Methode 2 : pivy coin
    try:
        from pivy import coin
        cam = view.getCameraNode()
        if cam:
            cam.position.setValue(cam_x, cam_y, cam_z)
            dx = cx - cam_x; dy = cy - cam_y; dz = cz - cam_z
            n = math.sqrt(dx*dx + dy*dy + dz*dz)
            if n > 0:
                ro = coin.SbRotation(coin.SbVec3f(0, 0, -1),
                                     coin.SbVec3f(dx/n, dy/n, dz/n))
                cam.orientation.setValue(ro)
            return True
    except: pass
    # Methode 3 : fallback isometric
    try:
        view.viewIsometric(); view.fitAll()
        return False
    except: return False

def _save_frame_safe(view, path, w, h):
    for attempt in range(3):
        try:
            view.saveImage(path, w, h, "White")
            if os.path.exists(path) and os.path.getsize(path) > 1000:
                return True
        except: pass
        time.sleep(0.02)
    return False

def _build_keyframes(cx, cy, cz, radius):
    kf = []
    for i in range(VID_FRAMES):
        t = (i / VID_FRAMES) * VID_DUR
        if t <= 3:
            e = _ease(t/3.0)
            dist=radius*(2.8-e*0.9); h=radius*(2.0-e*0.7); angle=math.radians(30)
        elif t <= 8:
            e = _ease((t-3)/5.0)
            dist=radius*(1.9-e*0.7); h=radius*(1.3-e*0.6); angle=math.radians(30+e*80)
        elif t <= 18:
            pct=(t-8)/10.0; angle=math.radians(110+pct*360)
            hwave=0.07*math.sin(pct*2*math.pi)
            dist=radius*1.05; h=radius*(0.42+hwave)
        elif t <= 23:
            e=_ease((t-18)/5.0)
            dist=radius*(1.05-e*0.65); h=radius*(0.42-e*0.15); angle=math.radians(470+e*40)
        elif t <= 27:
            e=_ease((t-23)/4.0)
            dist=radius*(0.4+e*0.65); h=radius*(0.27+e*0.20); angle=math.radians(510+e*30)
        else:
            e=_ease((t-27)/3.0)
            dist=radius*(1.05+e*0.6); h=radius*(0.47+e*1.0); angle=math.radians(540+e*20)
        kf.append((dist, h, angle))
    return kf

def module_video(doc, renders_dir):
    print("\n[3] Video Pro Full HD 1920x1080 (camera robuste)")
    try:
        view = FreeCADGui.activeDocument().activeView()
    except Exception as e:
        print(f"    [SKIP] Pas de vue active : {e}"); return None

    try: view.setAnimationEnabled(False)
    except: pass

    cx, cy, cz, radius = _scene_bounds(doc)
    print(f"    Centre : ({cx/1000:.1f}m, {cy/1000:.1f}m)  Rayon : {radius/1000:.1f}m")

    try: view.resize(VID_W, VID_H); _wait(400)
    except: pass

    _force_perspective(view)

    kf = _build_keyframes(cx, cy, cz, radius)
    frames_dir = os.path.join(renders_dir, "frames_tmp")
    os.makedirs(frames_dir, exist_ok=True)

    t0 = time.time()
    ok_count = fail_count = 0

    for i, (dist, h, angle) in enumerate(kf):
        _set_camera_robust(view, cx, cy, cz, dist, h, angle)
        try: QtCore.QCoreApplication.processEvents()
        except: pass

        fpath = os.path.join(frames_dir, f"frame_{i:05d}.png")
        if _save_frame_safe(view, fpath, VID_W, VID_H):
            ok_count += 1
        else:
            fail_count += 1
            # Copier la frame precedente pour eviter un trou dans la sequence
            if i > 0:
                prev = os.path.join(frames_dir, f"frame_{i-1:05d}.png")
                if os.path.exists(prev):
                    try:
                        import shutil; shutil.copy2(prev, fpath)
                    except: pass

        time.sleep(0.012)

        if i % 72 == 0 or i == VID_FRAMES - 1:
            elapsed = int(time.time() - t0)
            pct = (i + 1) / VID_FRAMES
            reste = int(elapsed / max(pct, 0.001) * (1 - pct))
            print(f"    [{i+1:3d}/{VID_FRAMES}] {int(pct*100)}%  "
                  f"OK:{ok_count} FAIL:{fail_count}  reste ~{min(reste,999)}s")

    print(f"\n    Frames : {ok_count} OK / {fail_count} echecs")

    mp4_name = f"{doc.Name}_presentation_investisseurs_HD.mp4"
    mp4_path = os.path.join(renders_dir, mp4_name)
    pattern  = os.path.join(frames_dir, "frame_%05d.png")

    ffmpeg_cmds = [
        ["ffmpeg", "-y", "-framerate", str(VID_FPS), "-i", pattern,
         "-c:v", "libx264", "-preset", "fast", "-crf", "20",
         "-pix_fmt", "yuv420p", "-vf", f"scale={VID_W}:{VID_H}", mp4_path],
        ["ffmpeg", "-y", "-framerate", str(VID_FPS), "-i", pattern,
         "-c:v", "libx265", "-preset", "fast", "-crf", "28",
         "-pix_fmt", "yuv420p", mp4_path],
        ["ffmpeg", "-y", "-framerate", str(VID_FPS), "-i", pattern,
         "-c:v", "mpeg4", "-q:v", "5", mp4_path],
    ]

    print("\n    Assemblage MP4...")
    video_ok = False
    for cmd in ffmpeg_cmds:
        try:
            r = subprocess.run(cmd, capture_output=True, timeout=600)
            if r.returncode == 0 and os.path.exists(mp4_path) and os.path.getsize(mp4_path) > 10000:
                sz = os.path.getsize(mp4_path) / 1e6
                print(f"    Video OK : {mp4_name} ({sz:.1f} MB)")
                video_ok = True; break
        except FileNotFoundError:
            print("    [WARN] ffmpeg non trouve (sudo apt install ffmpeg)"); break
        except subprocess.TimeoutExpired:
            print("    [WARN] ffmpeg timeout"); break
        except Exception as e:
            print(f"    [WARN] ffmpeg : {e}")

    try:
        import shutil; shutil.rmtree(frames_dir, ignore_errors=True)
    except: pass

    return mp4_path if video_ok else None

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 4 : CAPTURES 3D -- RETRY + FALLBACK
# ══════════════════════════════════════════════════════════════════════════════

CAPTURE_VIEWS = [
    ('01_isometric', lambda v: v.viewIsometric()),
    ('02_front',     lambda v: v.viewFront()),
    ('03_rear',      lambda v: v.viewRear()),
    ('04_top',       lambda v: v.viewTop()),
    ('05_bottom',    lambda v: v.viewBottom()),
    ('06_left',      lambda v: v.viewLeft()),
    ('07_right',     lambda v: v.viewRight()),
]

def module_captures(doc, renders_dir):
    print("\n[4] Captures 3D (7 vues avec retry)")
    captures = []
    try:
        view = FreeCADGui.activeDocument().activeView()
    except Exception as e:
        print(f"    [SKIP] {e}"); return captures

    for fname, set_view_fn in CAPTURE_VIEWS:
        path = os.path.join(renders_dir, f"{fname}.png")
        success = False
        for attempt in range(3):
            try:
                set_view_fn(view)
                _wait(250 + attempt * 100)
                view.fitAll()
                _wait(150)
                view.saveImage(path, VID_W, VID_H, "White")
                _wait(80)
                if os.path.exists(path) and os.path.getsize(path) > 500:
                    sz = os.path.getsize(path) // 1024
                    captures.append(path)
                    print(f"    OK {fname}.png ({sz} KB)")
                    success = True; break
            except Exception as e:
                if attempt == 2:
                    print(f"    [FAIL] {fname} : {e}")

        if not success and captures:
            import shutil
            shutil.copy2(captures[-1], path)
            captures.append(path)
            print(f"    [FALLBACK] {fname} copie depuis vue precedente")

    print(f"    {len(captures)}/{len(CAPTURE_VIEWS)} captures reussies")
    return captures

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 5 : COLLECTE DONNEES + MATERIAUX + AERODYNAMIQUE
# ══════════════════════════════════════════════════════════════════════════════

def module_collect(doc, fixes, wb_data):
    print("\n[5] Collecte donnees + materiaux IA + aerodynamique")
    fix_labels  = {f['label'] for f in fixes}
    objects     = []
    sketches    = []
    bodies      = []
    types_stats = {}
    masse_totale = 0.0
    aero_issues  = []

    for i, obj in enumerate(doc.Objects, 1):
        tid = obj.TypeId
        t   = tid.split('::')[-1]
        types_stats[t] = types_stats.get(t, 0) + 1

        if tid == 'PartDesign::Body':
            bodies.append({'Label': obj.Label}); continue

        if tid == 'Sketcher::SketchObject':
            geoms = 0; constrs = 0; detail = []
            try: geoms = len(obj.Geometry)
            except: pass
            try:
                constrs = len(obj.Constraints)
                for c in obj.Constraints[:15]:
                    detail.append({'Nom': str(c.Type),
                                   'Valeur': str(c.Value) if hasattr(c, 'Value') else ''})
            except: pass
            redundant = False
            try: redundant = obj.isValid() is False
            except: pass
            sketches.append({
                'Label': obj.Label, 'Geometries': geoms,
                'Contraintes': constrs, 'Detail': detail,
                'Statut': 'Redondant' if redundant else ('Contraint' if constrs > 0 else 'Libre'),
            }); continue

        fixed = obj.Label in fix_labels
        bb    = _bbox(obj)
        pos   = _get_pos(obj)
        vol   = _get_vol(obj)

        try: vis = 'Oui' if obj.Visibility else 'Non'
        except: vis = 'N/A'

        mat   = detect_material(obj)
        masse = calc_masse(obj, mat)
        if masse: masse_totale += masse

        aero = calc_aerodynamics(obj, bb)
        if aero and aero['Fd_N'] > 50000:
            aero_issues.append({
                'label': obj.Label,
                'Fd_N': aero['Fd_N'],
                'note': safe_text(f"Force vent critique {aero['Fd_N']:.0f} N - verifier ancrage"),
            })

        objects.append({
            '#': i,
            'Label': obj.Label,
            'Type': tid,
            'Visible': vis,
            'Materiau': mat['nom'],
            'Cat_Mat': mat['categorie'],
            'Densite': f"{mat['densite']} kg/m3" if mat['densite'] else 'N/A',
            'E_GPa': f"{mat['E_GPa']} GPa" if mat['E_GPa'] else 'N/A',
            'Re_MPa': f"{mat['Re_MPa']} MPa" if mat['Re_MPa'] else 'N/A',
            'Mat_Notes': safe_text(mat['notes']),
            'Masse_kg': masse,
            'Masse_str': f"{masse:.2f} kg" if masse else 'N/A',
            'Dimensions': _fmt_dims(bb),
            'Position': _fmt_pos(pos),
            'Rotation': f"rx={pos['rx']} ry={pos['ry']} rz={pos['rz']}",
            'Volume': _sv(vol) if vol else 'N/A',
            'Volume_raw': vol,
            'Aero_Re': f"{aero['Re']:.0f}" if aero else 'N/A',
            'Aero_Regime': aero['Re_regime'] if aero else 'N/A',
            'Aero_Cd': str(aero['Cd']) if aero else 'N/A',
            'Aero_Fd_N': f"{aero['Fd_N']:.1f} N" if aero else 'N/A',
            'Aero_M_Nm': f"{aero['M_Nm']:.1f} N.m" if aero else 'N/A',
            'Aero_P_Betz_W': f"{aero['P_betz_W']:.0f} W" if aero and aero['P_betz_W'] else 'N/A',
            'Statut': 'CORRIGE' if fixed else 'OK',
            '_fixed': fixed,
            '_aero': aero,
            '_mat': mat,
        })

    try:
        fn = doc.FileName
        date_mod = (datetime.datetime.fromtimestamp(os.path.getmtime(fn))
                    .strftime('%Y-%m-%d %H:%M') if fn else 'N/A')
    except: date_mod = 'N/A'

    general = {
        'Nom du document': doc.Name,
        'Fichier': doc.FileName or 'N/A',
        'Date du rapport': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'Derniere modif': date_mod,
        'Total objets': len(doc.Objects),
        'Corrections': len(fixes),
        'Masse totale est.': f"{masse_totale:.1f} kg" if masse_totale > 0 else 'N/A',
        'Script version': f"v{SCRIPT_VERSION}",
    }

    print("    Analyse IA du projet SmartFarm...")
    ai_analysis = analyze_project_ai(doc, objects)
    ai_analysis['aero_issues'] = aero_issues

    print(f"    {len(objects)} objets  |  masse ~{masse_totale:.0f} kg  |  "
          f"{ai_analysis['nb_systemes']} systemes detectes")

    return {
        'general': general,
        'objects': objects,
        'sketches': sketches,
        'bodies': bodies,
        'fixes': fixes,
        'types_stats': types_stats,
        'workbenches': wb_data,
        'ai_analysis': ai_analysis,
        'masse_totale': masse_totale,
    }

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT CSV
# ══════════════════════════════════════════════════════════════════════════════

def export_csv(data, path):
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['#','Label','Type','Visible','Materiau','Categorie',
                    'Densite','E_GPa','Re_MPa','Masse_kg',
                    'Dimensions','Position','Volume',
                    'Aero_Re','Aero_Regime','Aero_Cd','Aero_Fd_N','Aero_M_Nm','Aero_P_Betz_W',
                    'Statut','Mat_Notes'])
        for o in data['objects']:
            w.writerow([
                o['#'], o['Label'], o['Type'], o['Visible'],
                o['Materiau'], o['Cat_Mat'], o['Densite'], o['E_GPa'], o['Re_MPa'],
                o['Masse_kg'] if o['Masse_kg'] else '',
                o['Dimensions'], o['Position'], o['Volume'],
                o['Aero_Re'], o['Aero_Regime'], o['Aero_Cd'],
                o['Aero_Fd_N'], o['Aero_M_Nm'], o['Aero_P_Betz_W'],
                o['Statut'], o['Mat_Notes'],
            ])
    print(f"[CSV]  {path}")

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT JSON
# ══════════════════════════════════════════════════════════════════════════════

def export_json(data, path):
    def _clean(obj):
        if isinstance(obj, dict):
            return {k: _clean(v) for k, v in obj.items() if not k.startswith('_')}
        if isinstance(obj, list): return [_clean(i) for i in obj]
        if isinstance(obj, float) and not math.isfinite(obj): return None
        return obj

    payload = {
        'meta': {'script_version': SCRIPT_VERSION,
                 'generated': data['general']['Date du rapport'],
                 'project': data['general']['Nom du document']},
        'general': _clean(data['general']),
        'ai_analysis': _clean(data['ai_analysis']),
        'objects': _clean(data['objects']),
        'workbenches': {k: v for k, v in data['workbenches'].items() if v},
        'fixes': data['fixes'],
        'types_stats': data['types_stats'],
    }
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[JSON] {path}")

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT XLSX -- 6 onglets enrichis
# ══════════════════════════════════════════════════════════════════════════════

def _hc(ws, r, c, v, bg='2E74B5', fg='FFFFFF', bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fg, size=9, name='Arial')
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    th = Side(style='thin', color='AAAAAA')
    cell.border = Border(left=th, right=th, top=th, bottom=th)

def _dc(ws, r, c, v, bold=False, bg='FFFFFF', align='left', color='000000'):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='Arial', size=9, bold=bold, color=color)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    th = Side(style='thin', color='DDDDDD')
    cell.border = Border(left=th, right=th, top=th, bottom=th)

def _st_xl(ws, r, t, n, bg='1F3864'):
    c = ws.cell(row=r, column=1, value=f"  {t}")
    c.font = Font(bold=True, size=11, color='FFFFFF', name='Arial')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    ws.row_dimensions[r].height = 18

def export_xlsx(data, path):
    wb  = Workbook()
    nf  = len(data['fixes'])
    wbs = data.get('workbenches', {})
    ai  = data.get('ai_analysis', {})

    # ── Onglet 1 : Resume ────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = 'Resume'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 60

    ws1.merge_cells('A1:B1'); c = ws1['A1']
    c.value = f"DOSSIER TECHNIQUE IA - {data['general']['Nom du document']}"
    c.font  = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    c.fill  = PatternFill('solid', fgColor='1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:B2'); c2 = ws1['A2']
    c2.value = f"Genere le {data['general']['Date du rapport']} | Script v{SCRIPT_VERSION} | CONFIDENTIEL"
    c2.font  = Font(italic=True, size=9, color='444444', name='Arial')
    c2.fill  = PatternFill('solid', fgColor='DEEAF1')
    c2.alignment = Alignment(horizontal='center')

    row = 4
    _st_xl(ws1, row, 'INFORMATIONS GENERALES', 2); row += 1
    for k, v in data['general'].items():
        bg = 'F2F2F2' if row % 2 == 0 else 'FFFFFF'
        _dc(ws1, row, 1, k, bold=True, bg=bg); _dc(ws1, row, 2, str(v), bg=bg); row += 1

    row += 1; _st_xl(ws1, row, 'ANALYSE GLOBALE', 2); row += 1
    for k, v, grn in [
        ('Total objets', len(data['objects']), False),
        ('Corrections auto', nf, nf > 0),
        ('Sketches', len(data['sketches']), False),
        ('Bodies PartDesign', len(data['bodies']), False),
        ('Masse totale est.', data['general'].get('Masse totale est.', 'N/A'), False),
        ('Energie totale kW', f"{ai.get('total_energy_kW',0):.2f} kW", False),
        ('Systemes detectes', ai.get('nb_systemes', 0), False),
    ]:
        bg  = 'E2EFDA' if grn else ('F2F2F2' if row % 2 == 0 else 'FFFFFF')
        col = '375623' if grn else '000000'
        _dc(ws1, row, 1, k, bold=True, bg=bg)
        _dc(ws1, row, 2, str(v), bg=bg, align='center', color=col, bold=grn); row += 1

    row += 1; _st_xl(ws1, row, f'WORKBENCHES ({len(ALL_WB_LABELS)} scannes)', 2); row += 1
    for key, label in ALL_WB_LABELS:
        items = wbs.get(key)
        bg    = 'F2F2F2' if row % 2 == 0 else 'FFFFFF'
        if items:
            real = [it for it in items if it.get('type') != 'installed']
            status = f"{len(real)} element(s)" if real else 'Installe (0 objet)'
            col    = '375623' if real else 'B8860B'
        else:
            status = 'Non detecte'; col = '999999'
        _dc(ws1, row, 1, label, bold=bool(items), bg=bg)
        _dc(ws1, row, 2, status, bg=bg, align='center', color=col); row += 1

    # ── Onglet 2 : Analyse IA ────────────────────────────────────────────────
    ws_ai = wb.create_sheet('Analyse IA')
    ws_ai.sheet_view.showGridLines = False
    ws_ai.column_dimensions['A'].width = 32
    ws_ai.column_dimensions['B'].width = 90

    ws_ai.merge_cells('A1:B1'); c = ws_ai['A1']
    c.value = 'ANALYSE IA - INTERPRETATION SMARTFARM'
    c.font  = Font(bold=True, size=13, color='FFFFFF', name='Arial')
    c.fill  = PatternFill('solid', fgColor='375623')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_ai.row_dimensions[1].height = 26

    r = 3
    _st_xl(ws_ai, r, 'RAPPORT INTERPRETATIF', 2, '1F3864'); r += 1
    for para in ai.get('narrative', []):
        ws_ai.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws_ai.cell(row=r, column=1, value=para)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws_ai.row_dimensions[r].height = 45; r += 1

    r += 1; _st_xl(ws_ai, r, 'SYSTEMES FONCTIONNELS', 2, '2E74B5'); r += 1
    for sys_name, sys_data in ai.get('systems', {}).items():
        bg = 'DEEAF1' if r % 2 == 0 else 'FFFFFF'
        _dc(ws_ai, r, 1, sys_name, bold=True, bg='E2EFDA')
        _dc(ws_ai, r, 2, sys_data.get('note', ''), bg=bg); r += 1

    r += 1; _st_xl(ws_ai, r, 'RECOMMANDATIONS TECHNIQUES', 2, 'ED7D31'); r += 1
    for idx, rec in enumerate(ai.get('recommendations', []), 1):
        bg = 'FFF2CC' if r % 2 == 0 else 'FFFFFF'
        _dc(ws_ai, r, 1, f"R{idx:02d}", bold=True, bg=bg, align='center')
        _dc(ws_ai, r, 2, rec, bg=bg); r += 1

    if ai.get('aero_issues'):
        r += 1; _st_xl(ws_ai, r, 'ALERTES AERODYNAMIQUES', 2, 'C00000'); r += 1
        for issue in ai['aero_issues']:
            _dc(ws_ai, r, 1, issue['label'], bold=True, bg='FCE4D6')
            _dc(ws_ai, r, 2, issue['note'], bg='FCE4D6'); r += 1

    # ── Onglet 3 : Objets Enrichis ───────────────────────────────────────────
    ws2 = wb.create_sheet('Objets Enrichis')
    ws2.sheet_view.showGridLines = False
    hdrs   = ['#','Label','Type','Vis.','Materiau','Cat.','Densite','E(GPa)','Re(MPa)',
              'Masse(kg)','Dimensions','Volume','Aero Re','Regime','Cd','Fd(N)','M(N.m)','P Betz','Statut']
    widths = [4,26,16,6,24,12,10,8,8,10,42,12,12,11,6,9,9,10,9]
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        _hc(ws2, 1, i, h)
    for r, o in enumerate(data['objects'], 2):
        fx  = o['_fixed']
        bg  = 'E2EFDA' if fx else ('DEEAF1' if r % 2 == 0 else 'FFFFFF')
        sc  = '375623' if fx else '000000'
        mat_hex = str(o.get('_mat', {}).get('couleur', 'FFFFFF'))
        if len(mat_hex) != 6 or not all(c in '0123456789ABCDEFabcdef' for c in mat_hex):
            mat_hex = 'FFFFFF'
        vals = [o['#'], o['Label'], o['Type'].split('::')[-1], o['Visible'],
                o['Materiau'], o['Cat_Mat'], o['Densite'], o['E_GPa'], o['Re_MPa'],
                o['Masse_kg'], o['Dimensions'], o['Volume'],
                o['Aero_Re'], o['Aero_Regime'], o['Aero_Cd'],
                o['Aero_Fd_N'], o['Aero_M_Nm'], o['Aero_P_Betz_W'], o['Statut']]
        align_map = [1,0,0,1,0,1,1,1,1,1,0,1,1,1,1,1,1,1,1]
        for ci, (val, aln) in enumerate(zip(vals, align_map), 1):
            cell_bg = mat_hex if ci == 5 else bg
            _dc(ws2, r, ci, val, bold=(ci==2 or fx), bg=cell_bg,
                align='center' if aln else 'left', color=sc if fx else '000000')
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(data['objects'])+1}"

    # ── Onglet 4 : Materiaux Stats ───────────────────────────────────────────
    ws_mat = wb.create_sheet('Materiaux Stats')
    ws_mat.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(['Categorie','Nb objets','Masse totale kg','% masse','Exemple'],
                                    [28,10,16,10,50]), 1):
        ws_mat.column_dimensions[get_column_letter(i)].width = w
    _st_xl(ws_mat, 1, 'ANALYSE MATERIAUX PAR CATEGORIE', 5, '1F3864')
    for i, h in enumerate(['Categorie','Nb objets','Masse (kg)','% masse','Exemple materiau'], 1):
        _hc(ws_mat, 2, i, h)
    cat_stats = defaultdict(lambda: {'count': 0, 'masse': 0.0, 'examples': []})
    for o in data['objects']:
        cat = o.get('Cat_Mat', 'Inconnu')
        cat_stats[cat]['count'] += 1
        m = o.get('Masse_kg')
        if m: cat_stats[cat]['masse'] += m
        if len(cat_stats[cat]['examples']) < 2:
            cat_stats[cat]['examples'].append(o['Materiau'])
    mt = sum(v['masse'] for v in cat_stats.values()) or 1
    for r, (cat, info) in enumerate(sorted(cat_stats.items(), key=lambda x: -x[1]['masse']), 3):
        bg = 'F2F2F2' if r % 2 == 0 else 'FFFFFF'
        _dc(ws_mat, r, 1, cat, bold=True, bg=bg)
        _dc(ws_mat, r, 2, info['count'], bg=bg, align='center')
        _dc(ws_mat, r, 3, f"{info['masse']:.1f}", bg=bg, align='right')
        _dc(ws_mat, r, 4, f"{100*info['masse']/mt:.1f}%", bg=bg, align='center')
        _dc(ws_mat, r, 5, ' / '.join(set(info['examples'])), bg=bg)

    # ── Onglet 5 : Aerodynamique ─────────────────────────────────────────────
    ws_aero = wb.create_sheet('Aerodynamique')
    ws_aero.sheet_view.showGridLines = False
    aero_hdrs  = ['Label','Type','A_front(m2)','Re','Regime','Cd','Fd(N)','M(N.m)','P Betz(W)','Alerte']
    aero_widths = [28,14,12,14,12,6,10,10,11,32]
    for i, (h, w) in enumerate(zip(aero_hdrs, aero_widths), 1):
        ws_aero.column_dimensions[get_column_letter(i)].width = w
        _hc(ws_aero, 1, i, h, bg='15538A')
    ws_aero.merge_cells('A2:J2')
    c2 = ws_aero['A2']
    c2.value = f"Conditions : Vent ref = {V_WIND_REF} m/s ({V_WIND_REF*3.6:.0f} km/h) | RHO = {RHO_AIR} kg/m3"
    c2.font  = Font(italic=True, size=8, name='Arial')
    c2.fill  = PatternFill('solid', fgColor='EAF4FB')
    c2.alignment = Alignment(horizontal='center')
    r = 3
    for o in data['objects']:
        aero = o.get('_aero')
        if not aero: continue
        alerte = ''
        if aero['Fd_N'] > 100000: alerte = 'CRITIQUE - ancrage requis'
        elif aero['Fd_N'] > 50000: alerte = 'IMPORTANT - verifier'
        elif aero['Re_regime'] == 'Turbulent': alerte = 'Ecoulement turbulent'
        bg = 'FCE4D6' if 'CRITIQUE' in alerte else ('FFF2CC' if 'IMPORTANT' in alerte else
             ('FFFFFF' if r%2==0 else 'F2F2F2'))
        vals = [o['Label'], o['Type'].split('::')[-1],
                f"{aero['A_front_m2']:.4f}", f"{aero['Re']:.2e}",
                aero['Re_regime'], str(aero['Cd']),
                f"{aero['Fd_N']:.1f}", f"{aero['M_Nm']:.1f}",
                f"{aero['P_betz_W']:.0f}" if aero['P_betz_W'] else '-', alerte]
        for ci, v in enumerate(vals, 1):
            _dc(ws_aero, r, ci, v, bg=bg, bold=(ci==1),
                color='C00000' if 'CRITIQUE' in alerte else '000000')
        r += 1
    ws_aero.freeze_panes = 'A3'

    # ── Onglet 6 : Corrections ───────────────────────────────────────────────
    if data['fixes']:
        wsc = wb.create_sheet('Corrections')
        wsc.sheet_view.showGridLines = False
        wsc.merge_cells('A1:C1'); tc = wsc['A1']
        tc.value = f"{nf} correction(s)"
        tc.font  = Font(bold=True, size=11, color='FFFFFF', name='Arial')
        tc.fill  = PatternFill('solid', fgColor='375623')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        for i, (h, w) in enumerate(zip(['Label','Type','Actions'], [28,18,100]), 1):
            wsc.column_dimensions[get_column_letter(i)].width = w
            _hc(wsc, 2, i, h, bg='375623')
        for r, fx in enumerate(data['fixes'], 3):
            bg = 'E2EFDA' if r%2==0 else 'FFFFFF'
            _dc(wsc, r, 1, fx['label'], bold=True, bg=bg)
            _dc(wsc, r, 2, fx['type'], bg=bg)
            _dc(wsc, r, 3, ' | '.join(fx['actions']), bg=bg)

    # ── Onglets par WB detecte ───────────────────────────────────────────────
    for key, label in ALL_WB_LABELS:
        items = wbs.get(key)
        if not items: continue
        real = [it for it in items if it.get('type') != 'installed']
        if not real: continue
        sname = label[:31]
        if sname in [s.title for s in wb.worksheets]:
            sname = sname[:28] + str(sum(1 for s in wb.worksheets))
        ws = wb.create_sheet(sname)
        ws.sheet_view.showGridLines = False
        for i, (h, w) in enumerate(zip(['Label','Type'], [40,40]), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            _hc(ws, 1, i, h)
        for r, item in enumerate(real, 2):
            bg = 'DEEAF1' if r%2==0 else 'FFFFFF'
            _dc(ws, r, 1, item.get('label',''), bold=True, bg=bg)
            _dc(ws, r, 2, item.get('type',''), bg=bg)

    if data['sketches']:
        ws3 = wb.create_sheet('Sketches')
        ws3.sheet_view.showGridLines = False
        for i, (h, w) in enumerate(zip(['Label','Geom','Contraintes','Statut','Detail'],
                                        [28,8,12,14,65]), 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
            _hc(ws3, 1, i, h)
        for r, sk in enumerate(data['sketches'], 2):
            bg  = 'DEEAF1' if r%2==0 else 'FFFFFF'
            sbg = 'E2EFDA' if sk['Statut']=='Contraint' else 'FCE4D6'
            det = ' | '.join(f"{c['Nom']} {c['Valeur']}".strip() for c in sk['Detail']) or 'N/A'
            _dc(ws3, r, 1, sk['Label'], bold=True, bg=bg)
            _dc(ws3, r, 2, sk['Geometries'], bg=bg, align='center')
            _dc(ws3, r, 3, sk['Contraintes'], bg=bg, align='center')
            _dc(ws3, r, 4, sk['Statut'], bg=sbg, align='center')
            _dc(ws3, r, 5, det, bg=bg)

    wb.save(path)
    print(f"[XLSX] {path}")

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT PDF -- RAPPORT ENRICHI COMPLET
#  TOUTES les strings passent par safe_text() avant fpdf
# ══════════════════════════════════════════════════════════════════════════════

class PDF(FPDF):
    def __init__(self, pname, date):
        super().__init__()
        self.pname = safe_text(pname)
        self.date  = safe_text(date)
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_fill_color(31,56,100); self.rect(0,0,210,12,'F')
        self.set_font('Helvetica','B',9); self.set_text_color(255,255,255)
        self.set_xy(10,2)
        self.cell(0,8,st(f"DOSSIER TECHNIQUE IA  |  {self.pname}"),align='L')
        self.set_xy(0,2); self.cell(200,8,st(self.date),align='R')
        self.set_text_color(0,0,0); self.ln(12)

    def footer(self):
        self.set_y(-10); self.set_font('Helvetica','I',7)
        self.set_text_color(130,130,130)
        self.cell(0,5,st(f"Page {self.page_no()} - Dossier FreeCAD IA v{SCRIPT_VERSION} - CONFIDENTIEL"),align='C')

    def sec(self, t, col=(31,56,100)):
        self.ln(2); self.set_fill_color(*col); self.set_text_color(255,255,255)
        self.set_font('Helvetica','B',11)
        self.cell(0,8,st(f"  {t}"),fill=True,new_x='LMARGIN',new_y='NEXT')
        self.set_text_color(0,0,0); self.ln(2)

    def gbox(self, t):
        self.set_fill_color(226,239,218); self.set_draw_color(55,86,35)
        self.set_font('Helvetica','B',9); self.set_text_color(55,86,35)
        self.multi_cell(0,6,st(f"  {t}"),border=1,fill=True)
        self.set_text_color(0,0,0); self.set_draw_color(0,0,0); self.ln(2)

    def ibox(self, t):
        self.set_fill_color(222,234,241); self.set_draw_color(46,116,181)
        self.set_font('Helvetica','',9); self.set_text_color(31,56,100)
        self.multi_cell(0,6,st(f"  {t}"),border=1,fill=True)
        self.set_text_color(0,0,0); self.set_draw_color(0,0,0); self.ln(2)

    def warnbox(self, t):
        self.set_fill_color(255,235,156); self.set_draw_color(192,0,0)
        self.set_font('Helvetica','B',9); self.set_text_color(192,0,0)
        self.multi_cell(0,6,st(f"  ALERTE : {t}"),border=1,fill=True)
        self.set_text_color(0,0,0); self.set_draw_color(0,0,0); self.ln(2)

    def sub(self, t, col=(46,116,181)):
        self.set_text_color(*col); self.set_font('Helvetica','B',10)
        self.cell(0,6,st(t),new_x='LMARGIN',new_y='NEXT')
        self.set_text_color(0,0,0)

    def kv(self, d, w1=62):
        for i, (k, v) in enumerate(d.items()):
            self.set_fill_color(242,242,242) if i%2==0 else self.set_fill_color(255,255,255)
            self.set_font('Helvetica','B',9)
            self.cell(w1,6,st(str(k)),border=1,fill=True)
            self.set_font('Helvetica','',9)
            vs = st(str(v))[:90] + ('...' if len(st(str(v)))>90 else '')
            self.cell(0,6,vs,border=1,fill=True,new_x='LMARGIN',new_y='NEXT')
        self.ln(3)

    def tbl(self, headers, widths, rows, aligns=None):
        if aligns is None: aligns = ['C'] + ['L']*(len(headers)-1)
        self.set_fill_color(46,116,181); self.set_text_color(255,255,255)
        self.set_font('Helvetica','B',7)
        for h, w, a in zip(headers, widths, aligns):
            self.cell(w,7,st(str(h)),border=1,align=a,fill=True)
        self.ln(); self.set_text_color(0,0,0); self.set_font('Helvetica','',7)
        for i, row in enumerate(rows):
            flag = row[-1] if len(row) > len(widths) else ''
            if   'CORRIGE' in str(flag): self.set_fill_color(226,239,218)
            elif 'CRITIQUE' in str(flag): self.set_fill_color(255,200,180)
            elif i%2==0:                  self.set_fill_color(222,234,241)
            else:                         self.set_fill_color(255,255,255)
            for j, (val, w, a) in enumerate(zip(row[:len(widths)],widths,aligns)):
                vs = st(str(val) if val is not None else '')
                if len(vs) > 26 and w < 36: vs = vs[:23]+'...'
                grn = 'CORRIGE' in vs; red = 'CRITIQUE' in vs
                self.set_font('Helvetica','B' if (j==1 or grn or red) else '',7)
                if grn: self.set_text_color(55,86,35)
                if red: self.set_text_color(192,0,0)
                self.cell(w,6,vs,border=1,align=a,fill=True)
                if grn or red: self.set_text_color(0,0,0)
            self.ln()
            if self.get_y() > 268: self.add_page()
        self.ln(3)

    def narrative_block(self, paragraphs):
        self.set_font('Helvetica','',9)
        self.set_fill_color(248,250,252)
        for para in paragraphs:
            self.set_x(12)
            # safe_text deja applique dans analyze_project_ai()
            self.multi_cell(186,5.5,st(para),border='LB',fill=True)
            self.ln(1)
        self.ln(3)


def export_pdf(data, path, captures, video):
    g   = data['general']
    nf  = len(data['fixes'])
    ai  = data.get('ai_analysis', {})
    pdf = PDF(g['Nom du document'], g['Date du rapport'])
    pdf.add_page()

    # ── Cover ─────────────────────────────────────────────────────────────────
    pdf.set_fill_color(31,56,100); pdf.rect(10,20,190,40,'F')
    pdf.set_font('Helvetica','B',17); pdf.set_text_color(255,255,255)
    pdf.set_xy(10,24)
    pdf.cell(190,10,st('DOSSIER TECHNIQUE FREECAD - IA ENRICHI'),
             align='C',new_x='LMARGIN',new_y='NEXT')
    pdf.set_font('Helvetica','B',13); pdf.set_xy(10,36)
    pdf.cell(190,10,st(g['Nom du document']),align='C',new_x='LMARGIN',new_y='NEXT')
    pdf.set_font('Helvetica','',9); pdf.set_xy(10,49)
    pdf.cell(190,7,
             st(f"Rapport IA genere le {g['Date du rapport']} | Script v{SCRIPT_VERSION} | CONFIDENTIEL"),
             align='C',new_x='LMARGIN',new_y='NEXT')
    pdf.set_fill_color(237,125,49); pdf.rect(10,60,190,2,'F')
    pdf.set_text_color(0,0,0); pdf.ln(24)

    if nf > 0:
        pdf.gbox(st(f"{nf} correction(s) appliquee(s). Projet geometriquement coherent."))
    else:
        pdf.ibox(st("Projet OK - dimensions et positions coherentes. Aucune correction requise."))

    # 1. Infos generales
    pdf.sec(st("1. INFORMATIONS GENERALES"))
    pdf.kv(g)

    # 2. Statistiques
    pdf.sec(st("2. STATISTIQUES"))
    top = ', '.join(f"{t}:{c}" for t, c in
                    sorted(data['types_stats'].items(), key=lambda x: -x[1])[:6])
    pdf.kv({
        'Total objets':      len(data['objects']),
        'Corrections auto':  nf,
        'Types principaux':  st(top),
        'Sketches':          len(data['sketches']),
        'Bodies PartDesign': len(data['bodies']),
        'Masse totale est.': st(g.get('Masse totale est.', 'N/A')),
        'Energie installee': st(f"{ai.get('total_energy_kW',0):.1f} kW"),
        'Systemes detectes': ai.get('nb_systemes', 0),
    })

    # 3. Analyse IA
    pdf.sec(st("3. ANALYSE IA - INTERPRETATION DU PROJET"), (21,101,67))
    pdf.set_font('Helvetica','B',9); pdf.set_text_color(21,101,67)
    pdf.cell(0,6,st("  Analyse interpretative automatique du projet SmartFarm"),
             new_x='LMARGIN',new_y='NEXT')
    pdf.set_text_color(0,0,0)
    # narrative deja safe_text dans analyze_project_ai()
    pdf.narrative_block(ai.get('narrative', []))

    pdf.sub(st("  Systemes Fonctionnels Identifies :"), (21,101,67))
    for sys_name, sys_data in ai.get('systems', {}).items():
        pdf.set_font('Helvetica','B',8); pdf.set_fill_color(225,245,235)
        pdf.cell(52,5,st(f"  {sys_name}"),border=1,fill=True)
        pdf.set_font('Helvetica','',8); pdf.set_fill_color(248,255,250)
        note = st(sys_data.get('note',''))[:110]
        pdf.cell(0,5,note,border=1,fill=True,new_x='LMARGIN',new_y='NEXT')
    pdf.ln(3)

    if ai.get('aero_issues'):
        pdf.sub(st("  Alertes Aerodynamiques :"), (192,0,0))
        for iss in ai['aero_issues']:
            pdf.warnbox(st(f"{iss['label']} - {iss['note']}"))

    pdf.sub(st("  Recommandations Techniques :"))
    for idx, rec in enumerate(ai.get('recommendations', []), 1):
        pdf.set_font('Helvetica','',8); pdf.set_fill_color(255,248,220)
        pdf.set_x(12)
        pdf.cell(10,5,st(f"R{idx:02d}"),border=1,fill=True,align='C')
        pdf.set_fill_color(255,253,240)
        pdf.cell(0,5,st(rec),border=1,fill=True,new_x='LMARGIN',new_y='NEXT')
    pdf.ln(4)

    # 4. Workbenches
    wbs      = data.get('workbenches', {})
    wb_found = [k for k in wbs if wbs[k]]
    if wb_found:
        pdf.sec(st("4. WORKBENCHES DETECTES"), (15,80,60))
        wb_names = {k: lbl for k, lbl in ALL_WB_LABELS}
        for key in wb_found:
            items = wbs[key]
            real  = [it for it in items if it.get('type') != 'installed']
            lbl   = wb_names.get(key, key)
            if real:
                pdf.sub(st(f"  {lbl} ({len(real)} element(s))"))
                names = [it.get('label','?') for it in real[:6]]
                pdf.set_font('Helvetica','',8); pdf.set_text_color(60,60,60)
                pdf.cell(0,5,st('    '+', '.join(names)+('...' if len(real)>6 else '')),
                         new_x='LMARGIN',new_y='NEXT')
                pdf.set_text_color(0,0,0)
            else:
                pdf.set_font('Helvetica','I',8); pdf.set_text_color(150,150,0)
                pdf.cell(0,5,st(f"  {lbl} - addon installe (0 objet dans ce doc)"),
                         new_x='LMARGIN',new_y='NEXT')
                pdf.set_text_color(0,0,0)
            pdf.ln(1)

    sn = 5 if wb_found else 4

    # 5. Corrections
    if data['fixes']:
        pdf.sec(st(f"{sn}. CORRECTIONS APPLIQUEES"))
        pdf.tbl(['Label','Type','Corrections'],[40,22,125],
                [[st(f['label']),st(f['type']),st(' | '.join(f['actions'])),'CORRIGE']
                 for f in data['fixes']],['L','L','L'])
        sn += 1

    # 6. Liste objets
    pdf.sec(st(f"{sn}. LISTE COMPLETE DES OBJETS + MATERIAUX + AERODYNAMIQUE"))
    pdf.tbl(
        ['#','Label','Type','Materiau','Cat.','Masse','Dimensions','Vol.','Re','Cd','Fd(N)','Statut'],
        [6,26,14,28,10,12,36,11,13,6,9,9],
        [[o['#'],st(o['Label']),st(o['Type'].split('::')[-1]),
          st(o['Materiau']),st(o['Cat_Mat']),st(o['Masse_str']),
          st(o['Dimensions'][:33] if o['Dimensions']!='N/A' else 'N/A'),
          st(o['Volume']),st(o['Aero_Re']),st(o['Aero_Cd']),
          st(o['Aero_Fd_N'].replace(' N','')),st(o['Statut']),st(o['Statut'])]
         for o in data['objects']],
        ['C','L','L','L','C','R','L','R','R','C','R','C'])
    sn += 1

    # 7. Sketches
    if data['sketches']:
        pdf.sec(st(f"{sn}. SKETCHES"))
        for sk in data['sketches']:
            pdf.sub(st(f"  {sk['Label']}"))
            pdf.kv({'Geometries': sk['Geometries'],
                    'Contraintes': sk['Contraintes'],
                    'Statut': st(sk['Statut'])})
        sn += 1

    # 8. Livrables 3D
    if captures or video:
        pdf.sec(st(f"{sn}. LIVRABLES 3D"))
        if captures:
            pdf.sub(st(f"  Captures PNG ({len(captures)} x {VID_W}x{VID_H})"))
            for c in captures[:7]:
                pdf.set_font('Helvetica','',8); pdf.set_text_color(60,60,60)
                pdf.cell(0,5,st(f"    {os.path.basename(c)}"),new_x='LMARGIN',new_y='NEXT')
            pdf.set_text_color(0,0,0)
        if video:
            pdf.ln(2); pdf.sub(st("  Video Presentation Investisseurs Full HD"))
            pdf.set_font('Helvetica','',9)
            for line in [
                st(f"    Fichier : {os.path.basename(video)}"),
                st(f"    Format  : MP4 H264 | {VID_W}x{VID_H} | {VID_FPS}fps | {VID_DUR}s"),
                st(f"    Contenu : Intro aerienne + Rotation 360 + Zoom technique"),
            ]:
                pdf.cell(0,6,line,new_x='LMARGIN',new_y='NEXT')

    pdf.output(path)
    print(f"[PDF]  {path}")

# ══════════════════════════════════════════════════════════════════════════════
#  MODULE 10 : BREVET INNORPI -- REMPLISSAGE AUTOMATIQUE
#  Genere un formulaire INNORPI pre-rempli avec les infos du projet
#  Requiert : pypdf + reportlab (pip install pypdf reportlab)
# ══════════════════════════════════════════════════════════════════════════════

BREVET_DATA = {
    'title_line1': "Integrated Autonomous Smart Farming Ecosystem",
    'title_line2': "for Livestock and Greenhouse Management",
    'nom_prenom':  "Mhiri Ahmed",
    'nationalite': "Tunisienne",
    'rue':         "Route el Ain km 4.5",
    'cp_ville':    "3042 Sfax",
    'pays':        "Tunisie",
    'tel':         "+216 29 431 551",
    'email':       "mhiriahmed478@gmail.com",
    'inv1_nom':    "Mhiri Ahmed",
    'inv1_adr':    "Route el Ain km 4.5, 3042 Sfax, Tunisie",
    'signataire':  "Mhiri Ahmed",
}

# Chemin du formulaire vierge -- a adapter selon votre machine
BREVET_TEMPLATE = os.path.expanduser(
    "~/REQUETE_DE_DEPOT_-_Ar_Fr__En_-_Copie.pdf"
)

def brevet(template_path=None, output_dir=None):
    """
    Remplit automatiquement le formulaire INNORPI de depot de brevet.
    Usage : brevet()
    """
    if not RL_OK:
        print("[BREVET] reportlab non installe.")
        print("         pip install reportlab pypdf --break-system-packages")
        return None

    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        print("[BREVET] pypdf non installe. pip install pypdf --break-system-packages")
        return None

    tpl = template_path or BREVET_TEMPLATE
    if not os.path.exists(tpl):
        # Chercher dans les repertoires courants
        for candidate in [
            os.path.expanduser("~/REQUETE_DE_DEPOT_-_Ar_Fr__En_-_Copie.pdf"),
            os.path.expanduser("~/Documents/REQUETE_DE_DEPOT_-_Ar_Fr__En_-_Copie.pdf"),
            "/tmp/REQUETE_DE_DEPOT_-_Ar_Fr__En_-_Copie.pdf",
        ]:
            if os.path.exists(candidate):
                tpl = candidate; break
        else:
            print(f"[BREVET] Template non trouve : {tpl}")
            print("         Copiez le formulaire INNORPI vierge dans ~/")
            return None

    out_dir = output_dir or os.path.dirname(tpl)
    out_path = os.path.join(out_dir, "INNORPI_Brevet_Ahmed_Mhiri.pdf")

    print("\n[BREVET] Generation formulaire INNORPI...")

    PAGE_H = 841.9

    def ry(struct_y, font_size=9):
        return PAGE_H - struct_y - font_size

    def draw_x(c, cx, cy_struct, size=8):
        rl_y = PAGE_H - cy_struct - size / 2
        c.setFont("Helvetica-Bold", size)
        c.drawCentredString(cx, rl_y, "X")

    def overlay_page1(c):
        c.setFillColorRGB(0,0,0)
        draw_x(c, 370.6, 91.4, 8)         # Langue Francais
        c.setFont("Helvetica-Bold", 9)
        c.drawString(36, ry(332, 9), BREVET_DATA['title_line1'])
        c.setFont("Helvetica", 8.5)
        c.drawString(36, ry(346, 8.5), BREVET_DATA['title_line2'])

    def overlay_page2(c):
        c.setFillColorRGB(0,0,0)
        draw_x(c, 559, 167, 9)            # Personne physique
        c.setFont("Helvetica", 9)
        c.drawString(375, ry(208, 9), BREVET_DATA['nom_prenom'])
        c.drawString(90,  ry(262, 9), BREVET_DATA['nationalite'])
        c.drawString(90,  ry(281, 9), BREVET_DATA['rue'])
        c.drawString(90,  ry(301, 9), BREVET_DATA['cp_ville'])
        c.drawString(90,  ry(321, 9), BREVET_DATA['pays'])
        c.drawString(90,  ry(352, 9), BREVET_DATA['tel'])
        c.drawString(380, ry(352, 9), BREVET_DATA['email'])
        c.drawString(115, ry(427, 9), BREVET_DATA['inv1_nom'])
        c.drawString(115, ry(447, 9), BREVET_DATA['inv1_adr'])

    def overlay_page3(c):
        c.setFillColorRGB(0,0,0)
        draw_x(c, 279.9, 215.7, 8)       # Description
        draw_x(c, 279.9, 249.4, 8)       # Revendications
        draw_x(c, 279.9, 283.0, 8)       # Abrege
        c.setFont("Helvetica", 9)
        c.drawString(410, ry(224, 9), "3")
        c.drawString(410, ry(272, 9), "1")
        c.drawString(410, ry(339, 9), "5")
        c.drawString(130, ry(452, 9), BREVET_DATA['signataire'])
        draw_x(c, 123.7, 517.6, 8)       # Deposant
        draw_x(c, 341.9, 664.3, 8)       # Bureau Sfax

    def make_overlay(fn):
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        fn(c); c.save(); buf.seek(0)
        return buf

    # Remplir les champs fillable
    field_values = [
        {"field_id": "Case \u00e0 cocher3", "value": "/Oui"},   # Francais
        {"field_id": "Case \u00e0 cocher2", "value": "/Off"},
        {"field_id": "Case \u00e0 cocher1", "value": "/Off"},
        {"field_id": "Texte1",              "value": ""},
        {"field_id": "Case \u00e0 cocher4", "value": "/Off"},
        {"field_id": "Texte2",              "value": ""},
        {"field_id": "Texte5",              "value": ""},
        {"field_id": "Texte3",              "value": ""},
    ]

    try:
        reader = PdfReader(tpl)
        writer = PdfWriter()

        # Remplir les champs AcroForm
        for page in reader.pages:
            writer.add_page(page)
        for fv in field_values:
            try:
                writer.update_page_form_field_values(
                    writer.pages[0] if fv.get('page',1)==1 else writer.pages[1],
                    {fv['field_id']: fv['value']}
                )
            except: pass

        # Ecrire base dans buffer
        base_buf = io.BytesIO()
        writer.write(base_buf)
        base_buf.seek(0)

        # Overlay textes
        reader2  = PdfReader(base_buf)
        writer2  = PdfWriter()
        overlays = {0: overlay_page1, 1: overlay_page2, 2: overlay_page3}
        for i, page in enumerate(reader2.pages):
            if i in overlays:
                ov_buf = make_overlay(overlays[i])
                ov_pg  = PdfReader(ov_buf).pages[0]
                page.merge_page(ov_pg)
            writer2.add_page(page)

        with open(out_path, "wb") as f:
            writer2.write(f)

        print(f"[BREVET] Formulaire genere : {out_path}")
        return out_path

    except Exception as e:
        print(f"[BREVET] Erreur : {e}")
        import traceback; traceback.print_exc()
        return None

# ══════════════════════════════════════════════════════════════════════════════
#  CHATBOT / AGENT INTERACTIF
# ══════════════════════════════════════════════════════════════════════════════

_CHATBOT_DATA = {}

def _chatbot_answer(question, data):
    q = question.lower().strip()

    if any(k in q for k in ('energie','energy','kwh','kw','solaire','puissance')):
        ai = data.get('ai_analysis', {})
        sys = ai.get('systems', {})
        lines = [f"Energie totale installee : {ai.get('total_energy_kW',0):.2f} kW"]
        for name, info in sys.items():
            lines.append(f"  - {name} : {info.get('note','')}")
        return '\n'.join(lines)

    if any(k in q for k in ('masse','mass','poids','weight','kg')):
        mt = data.get('masse_totale', 0)
        return f"Masse totale estimee : {mt:.1f} kg ({mt/1000:.2f} t)"

    if any(k in q for k in ('materia','material','acier','beton','aluminium')):
        cats = defaultdict(int)
        for o in data.get('objects', []):
            cats[o.get('Cat_Mat','?')] += 1
        lines = ["Categories materiaux :"]
        for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
            lines.append(f"  - {cat} : {n} objet(s)")
        return '\n'.join(lines)

    if any(k in q for k in ('aero','vent','wind','reynolds','trainee','drag')):
        objs = data.get('objects', [])
        max_fd = max((o.get('_aero',{}).get('Fd_N',0) or 0 for o in objs), default=0)
        turb   = [o['Label'] for o in objs if o.get('_aero',{}).get('Re_regime')=='Turbulent']
        betz   = [o for o in objs if o.get('Aero_P_Betz_W','N/A')!='N/A']
        lines  = [
            f"Force de trainee max : {max_fd:.0f} N (vent {V_WIND_REF} m/s)",
            f"Objets en regime turbulent : {len(turb)}",
            f"Turbines avec P Betz : {len(betz)}",
        ]
        for o in betz[:3]:
            lines.append(f"  - {o['Label']} : {o['Aero_P_Betz_W']}")
        return '\n'.join(lines)

    if any(k in q for k in ('workbench','wb','addon','plugin')):
        wbs = data.get('workbenches', {})
        found = [(k,v) for k,v in wbs.items() if v and any(it.get('type')!='installed' for it in v)]
        lines = [f"Workbenches avec objets : {len(found)}"]
        for k, v in found[:8]:
            real = [it for it in v if it.get('type')!='installed']
            lines.append(f"  - {k} : {len(real)} element(s)")
        return '\n'.join(lines)

    if any(k in q for k in ('objet','object','combien','total','count')):
        objs = data.get('objects', [])
        return (f"Total : {len(objs)} objets\n"
                "Types : " + ', '.join(f"{t}:{c}" for t,c in
                sorted(data.get('types_stats',{}).items(),key=lambda x:-x[1])[:5]))

    if any(k in q for k in ('recommand','conseil','advice','improve')):
        recs = data.get('ai_analysis',{}).get('recommendations',[])
        if not recs: return "Aucune recommandation."
        return "Recommandations :\n" + '\n'.join(f"  R{i:02d}. {r}" for i,r in enumerate(recs,1))

    if any(k in q for k in ('systeme','system','hydro','turbine','ruche')):
        sys = data.get('ai_analysis',{}).get('systems',{})
        lines = [f"Systemes detectes ({len(sys)}) :"]
        for name, info in sys.items():
            lines.append(f"  - {name} : {str(info.get('note',''))[:80]}")
        return '\n'.join(lines)

    if any(k in q for k in ('fix','correct','correction','repare')):
        fixes = data.get('fixes',[])
        if not fixes: return "Aucune correction. Projet geometriquement correct."
        return (f"{len(fixes)} correction(s) :\n" +
                '\n'.join(f"  - {f['label']} : {' | '.join(f['actions'])}" for f in fixes[:10]))

    if any(k in q for k in ('brevet','innorpi','patent','formulaire')):
        return ("Tapez brevet() pour generer le formulaire INNORPI pre-rempli.\n"
                "Le formulaire sera sauvegarde dans votre dossier home.")

    return ("Question non reconnue. Sujets disponibles :\n"
            "  energie | masse | materiaux | aerodyn | workbench\n"
            "  objets  | systemes | recommand | brevet | aide")

def chatbot(data=None):
    d = data or _CHATBOT_DATA
    if not d:
        print("[CHATBOT] Lancez d'abord run() pour charger les donnees.")
        return
    print("\n" + "="*56)
    print(f"  CHATBOT SMARTFARM IA v{SCRIPT_VERSION}")
    print("  Posez vos questions (tapez 'aide' ou 'quit')")
    print("="*56)
    while True:
        try:
            q = input("\n  Question > ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n  [Chatbot ferme]"); break
        if not q: continue
        if q.lower() in ('quit','exit','q','bye'):
            print("  [Chatbot ferme]"); break
        answer = _chatbot_answer(q, d)
        print("\n" + "-"*50)
        print(answer)
        print("-"*50)

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def run():
    print("\n" + "="*60)
    print(f"  FreeCAD Master Script v{SCRIPT_VERSION}")
    print("  IA Materiaux + Aerodynamique + Analyse Projet")
    print("  Fix + Workbenches + Video + Captures + Rapport + Brevet")
    print("="*60)

    if not FREECAD:
        print("[ERREUR] FreeCAD non disponible."); return

    doc = FreeCAD.ActiveDocument
    if not doc: raise RuntimeError("Aucun document ouvert.")

    proj, base_dir, renders_dir = _out_dir(doc)

    # 1. Fix geometrique
    fixes = module_fix(doc)
    if fixes:
        try: doc.save(); print("    Sauvegarde OK")
        except: pass

    # 2. Scan workbenches
    wb_data = module_workbenches(doc)

    # 3. Video Full HD
    video = module_video(doc, renders_dir)

    # 4. Captures 7 vues
    captures = module_captures(doc, renders_dir)

    # 5. Collecte + IA
    data = module_collect(doc, fixes, wb_data)

    # 6. Exports
    print("\n[6] Generation rapports")
    csv_path  = os.path.join(base_dir, f"{proj}_rapport.csv")
    xlsx_path = os.path.join(base_dir, f"{proj}_rapport.xlsx")
    pdf_path  = os.path.join(base_dir, f"{proj}_rapport.pdf")
    json_path = os.path.join(base_dir, f"{proj}_rapport.json")

    export_csv(data,  csv_path)
    export_json(data, json_path)
    export_xlsx(data, xlsx_path)
    export_pdf(data,  pdf_path, captures, video)

    # Rendre les donnees disponibles pour chatbot()
    global _CHATBOT_DATA
    _CHATBOT_DATA.update(data)

    ai = data.get('ai_analysis', {})

    print("\n" + "="*60)
    print(f"  DONE - FreeCAD Master Script v{SCRIPT_VERSION}")
    print(f"  Rapports  : {base_dir}")
    print(f"  Renders   : {renders_dir}")
    print(f"  Objets    : {len(data['objects'])}")
    print(f"  Captures  : {len(captures)} PNG")
    if video: print(f"  Video     : {os.path.basename(video)}")
    if fixes: print(f"  Fix       : {len(fixes)} correction(s)")
    print(f"  Systemes  : {ai.get('nb_systemes',0)} detectes")
    print(f"  Energie   : {ai.get('total_energy_kW',0):.1f} kW estimes")
    print(f"  Masse     : ~{data.get('masse_totale',0):.0f} kg")
    if ai.get('aero_issues'):
        print(f"  ALERTES   : {len(ai['aero_issues'])} alerte(s) aerodynamiques !")
    print(f"\n  Fichiers :")
    for f in [pdf_path, xlsx_path, csv_path, json_path]:
        print(f"    {os.path.basename(f)}")
    print(f"\n  COMMANDES DISPONIBLES APRES run() :")
    print(f"    chatbot()   -> Q&R interactif sur le projet")
    print(f"    brevet()    -> Formulaire INNORPI pre-rempli")
    print("="*60 + "\n")

    return data


try:
    result = run()
except RuntimeError as e:
    print(f"\n[ERREUR] {e}\n")
except Exception:
    import traceback; traceback.print_exc()
